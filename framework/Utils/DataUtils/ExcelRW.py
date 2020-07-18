import openpyxl
import time

class ExcelReader(object):

    def get_data_by_header_and_tc_id(self, logger, file_path=None, sheet_name=None, tc_id=None, header_name=None):
        try:
            self.logger = logger
            book = openpyxl.load_workbook(filename=file_path)
            sheet = book.get_sheet_by_name(sheet_name)
            tc_id = tc_id
            header_name = header_name
            row_no = None
            column_no = None

            headers = [cell.value for cell in next(sheet.rows)]
            for header in headers:
                if header == header_name:
                    # get column no
                    column_no = headers.index(header)+1
                    break

            for col in sheet.iter_cols(1, sheet.max_column):
                for cell in col:
                    # print(cell.value)
                    if cell.value == tc_id:
                        #get col no
                        row_no = cell.row
                        break
            # if row_no or column_no is None:
            #     logger.error('ExcelReader: Unable to find correct cell in Excel File!')
            #     return False
            data = sheet.cell(row=row_no, column=column_no).value
            return data
        except Exception as e:
            self.logger.error(e)
            return False
    
    def get_column_no_by_header(self, logger, file_path=None, sheet_name=None, header_name=None):
        try:
            self.logger = logger
            book = openpyxl.load_workbook(filename=file_path)
            sheet = book.get_sheet_by_name(sheet_name)
            header_name = header_name

            headers = [cell.value for cell in next(sheet.rows)]
            for header in headers:
                if header == header_name:
                    # get column no
                    column_no = headers.index(header)+1
                    break
            # if column_no is None:
            #     logger.error('ExcelReader: Unable to find correct cell in Excel File!')
            #     return False
            return column_no
        except Exception as e:
            self.logger.error(e)
            return False

    def get_row_no_by_tc_id(self, logger, file_path=None, sheet_name=None, tc_id=None):
        try:
            self.logger = logger
            book = openpyxl.load_workbook(filename=file_path)
            sheet = book.get_sheet_by_name(sheet_name)
            tc_id = tc_id

            for col in sheet.iter_cols(1, sheet.max_column):
                for cell in col:
                    # print(cell.value)
                    if cell.value == tc_id:
                        #get col no
                        row_no = cell.row
                        break
            # if row_no is None:
            #     logger.error('ExcelReader: Unable to find correct cell in Excel File!')
            #     return False
            return row_no
        except Exception as e:
            self.logger.error(e)
            return False

class ExcelWriter:
    def set_test_result(self, logger, file_path=None, sheet_name=None, testcase_id=None, header_name=None, result=None):
        try:
            self.logger = logger
            book = openpyxl.load_workbook(filename=file_path)
            sheet = book.get_sheet_by_name(sheet_name)
            tc_id = testcase_id
            header_name = header_name
            row_no = None
            column_no = None

            headers = [cell.value for cell in next(sheet.rows)]
            for header in headers:
                if header == header_name:
                    # get column no
                    column_no = headers.index(header)+1
                    break

            for col in sheet.iter_cols(1, sheet.max_column):
                for cell in col:
                    # print(cell.value)
                    if cell.value == tc_id:
                        #get col no
                        row_no = cell.row
                        break
            self.logger.info("write: %s %s"%(row_no, column_no))
            # if row_no or column_no is None:
            #     logger.error('ExcelWriter: Unable to find correct cell in Excel File!')
            #     return False
            sheet.cell(row=row_no, column=column_no, value=result)
            book.save(file_path)
            logger.info("Set Result [%s] for Testcase [%s] in file: %s"%(result, tc_id, file_path))

            return True
        except Exception as e:
            self.logger.error(e)
            return False
# ExcelRead.get_data_by_header_and_tc_id()